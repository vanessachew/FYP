import nltk
from rake_nltk import Rake
# import pandas as pd 
from nltk.tokenize import sent_tokenize
# import torch
from rouge import Rouge
from openpyxl import load_workbook
from csv import DictWriter

# Download necessary NLTK resources
nltk.download('punkt')
nltk.download('stopwords')

# Open the workbook
workbook = load_workbook('output.xlsx')

outfile_path = "rouge_scores_word_count.xlsx"

rouge = Rouge()

# Specify the rows and columns you want to read
start_row = 2  # Start from row 2
end_row = 55    # End at row 55
start_col = 0  # Start from column B (assuming the first column is A)
end_col = 25    # End at column G

# Access the active worksheet
worksheet = workbook.active

# Iterate over each row in the specified range
for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=51, min_col=1, max_col=25, values_only=True), start=2):
    # Accessing specific columns
    text = row[1]  # Assuming column B contains the text (index 0 because indexing starts from 0)
    summary = row[3]  # Assuming column D contains the summary (index 2)
    summary_gpt = row[14]  # Assuming column F contains the GPT summary (index 4)

    # calc word count
    word_count_text = len(text)
    word_count_summ = len(summary)
    word_count_gpt = len(summary_gpt)

    # Update the Excel file with the word count for each column
    worksheet.cell(row=row_num, column=3, value=word_count_text)  # Assuming column G for text word count
    worksheet.cell(row=row_num, column=5, value=word_count_summ)  # Assuming column H for summary word count
    worksheet.cell(row=row_num, column=16, value=word_count_gpt)  # Assuming column I for GPT summary word count

    # # Save the updated workbook
    # workbook.save('output_with_word_count.xlsx')

    # scores_text = rouge.get_scores(text, summary, avg=True)    
    #     print(scores_text)
    #     print(f'Length of RRR summary:{word_count_summ}')

    # scores_gpt = rouge.get_scores(text, summary_gpt, avg=True)    
    #     print(scores_gpt)
    #     print(f'Length of GPT summary:{word_count_gpt}')

# Calculate ROUGE scores for summary and GPT summary
    scores_summary = rouge.get_scores(text, summary, avg=True)
    scores_summary_gpt = rouge.get_scores(text, summary_gpt, avg=True)

    # Update the Excel file with ROUGE scores
    worksheet.cell(row=row_num, column=6, value=scores_summary['rouge-1']['r'])
    worksheet.cell(row=row_num, column=7, value=scores_summary['rouge-1']['p'])
    worksheet.cell(row=row_num, column=8, value=scores_summary['rouge-1']['f'])
    worksheet.cell(row=row_num, column=9, value=scores_summary['rouge-2']['r'])
    worksheet.cell(row=row_num, column=10, value=scores_summary['rouge-2']['p'])
    worksheet.cell(row=row_num, column=11, value=scores_summary['rouge-2']['f'])
    worksheet.cell(row=row_num, column=12, value=scores_summary['rouge-l']['r'])
    worksheet.cell(row=row_num, column=13, value=scores_summary['rouge-l']['p'])
    worksheet.cell(row=row_num, column=14, value=scores_summary['rouge-l']['f'])
    worksheet.cell(row=row_num, column=17, value=scores_summary_gpt['rouge-1']['r'])
    worksheet.cell(row=row_num, column=18, value=scores_summary_gpt['rouge-1']['p'])
    worksheet.cell(row=row_num, column=19, value=scores_summary_gpt['rouge-1']['f'])
    worksheet.cell(row=row_num, column=20, value=scores_summary_gpt['rouge-2']['r'])
    worksheet.cell(row=row_num, column=21, value=scores_summary_gpt['rouge-2']['p'])
    worksheet.cell(row=row_num, column=22, value=scores_summary_gpt['rouge-2']['f'])
    worksheet.cell(row=row_num, column=23, value=scores_summary_gpt['rouge-l']['r'])
    worksheet.cell(row=row_num, column=24, value=scores_summary_gpt['rouge-l']['p'])
    worksheet.cell(row=row_num, column=25, value=scores_summary_gpt['rouge-l']['f'])

    # # Print out the ROUGE scores for both summaries
    # print(f"Row {row_num}:")
    # print(f"Text summary score: {scores_summary}")
    # print(f"GPT summary score: {scores_summary_gpt}")

    # # Storing ROUGE scores in the columns below the current columns
    # for metric in ['rouge-1', 'rouge-2', 'rouge-l']:
    #     for aspect in ['r', 'p', 'f']:
    #         worksheet.cell(row=row_num, column=3, value=scores_summary[metric][aspect])
    #         worksheet.cell(row=row_num, column=4, value=scores_summary_gpt[metric][aspect])


# Save the updated workbook
workbook.save(outfile_path)



# # Iterate over each row in the specified range
# for row_num, row in enumerate(worksheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=True), start=start_row):
#     # Accessing specific columns
#     text = row[1]  # Assuming column B contains the text (index 0 because indexing starts from 0)
#     summary = row[3]  # Assuming column D contains the summary (index 2)
#     summary_gpt = row[5]  # Assuming column F contains the GPT summary (index 4)

#     # Calculate ROUGE scores for summary and GPT summary
#     scores_summary = rouge.get_scores(text, summary, avg=True)
#     scores_summary_gpt = rouge.get_scores(text, summary_gpt, avg=True)

#     # Update the Excel file with ROUGE scores
#     field_names = ['rouge-1', 'rouge-2', 'rouge-l']
#     field_names_x = ['f', 'p', 'r']
#     with open(outfile_path + 'output.xlsx','w') as outfile:
#         writer = DictWriter(outfile, fieldnames=field_names)
#         writer.writeheader()
#         writer.writerows(scores_summary)

# # Save the updated workbook
# # workbook.save('output_with_word_count.xlsx')